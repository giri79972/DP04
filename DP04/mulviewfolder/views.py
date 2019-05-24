from django.shortcuts import render
from django.http.response import HttpResponse
import os
import PyPDF2
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from PyPDF2 import PdfFileReader
from time import time


# Create your views here.
def home(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "PDF Files MataData"
    #excel label name
    ws.cell(1,1,value='number_of_pages')
    ws.cell(1,2,value='info.author')
    ws.cell(1,3,value='info.creator')
    ws.cell(1,4,value='info.producer')
    ws.cell(1,5,value='info.subject')
    ws.cell(1,6,value='info.title')
    ws.cell(1,7,value='Path')
    #excel file
    def excel(i,j,value):
        if(j==1):
            ws.cell(row=i, column=j,value="{0}".format(value))
        if(j==2):
            ws.cell(row=i, column=j,value="{0}".format(value))
        if(j==3):
            ws.cell(row=i, column=j,value="{0}".format(value))
        if(j==4):
            ws.cell(row=i, column=j,value="{0}".format(value))
        if(j==5):
            ws.cell(row=i, column=j,value="{0}".format(value))
        if(j==6):
            ws.cell(row=i, column=j,value="{0}".format(value))
        if(j==7):
            ws.cell(row=i, column=j,value="{0}".format(value))
    #file operation dir and sub-dir checking process
    def file02():
        i=1
        for r, d, f in os.walk('D://Backup//16122018//23112018//Raspberry Pi//'):
            #i=i+1
            for file in f:
                #i=i+1
                try:
                    #i=0
                    if ".pdf" in file:
                        i=i+1
                        with open(os.path.join(r, file), 'rb') as f:
                            try:
                                print(os.path.join(r, file))
                                j=0
                                pdf = PdfFileReader(f)
                                info = pdf.getDocumentInfo()
                                number_of_pages = pdf.getNumPages()
                                print(number_of_pages)
                                j=j+1
                                excel(i,j,number_of_pages)
                                print(info.author)
                                j=j+1
                                excel(i,j,info.author)
                                print(info.creator)
                                j=j+1
                                excel(i,j,info.creator)
                                print(info.producer)
                                j=j+1
                                excel(i,j,info.producer)
                                print(info.subject)
                                j=j+1
                                excel(i,j,info.subject)
                                print(info.title)
                                j=j+1
                                excel(i,j,info.title)
                                j=j+1
                                excel(i,j,os.path.join(r, file))
                            except PyPDF2.utils.PdfReadError:
                                print("PyPDF2.utils.PdfReadError")
                            print("Completed")
                            #count=count+1
                except Exception as e:
                    i=i-1
                    print(e)
    file02() #file count is ok
    #print(file02())
    #excel()
    wb.save('pdfmatadatalog.xlsx')
    return HttpResponse(file02())

def contact(request):
    c="<h1>Contact Page Logic</h1>"
    return HttpResponse(c)

def services(request):
    s="<h1>Services Page Logic</h1>"
    return HttpResponse(s)
def feedback(request):
    f="<h1>Feedback Page Logic</h1>"
    return HttpResponse(f)